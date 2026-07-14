from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QHeaderView,
    QAbstractItemView, QDialog, QFormLayout, QStackedWidget,
    QTableWidgetItem
)
from PyQt6.QtCore import Qt
from qfluentwidgets import (
    TitleLabel, SubtitleLabel, TableWidget, PushButton, PrimaryPushButton,
    LineEdit, MessageBox, FluentIcon, BodyLabel, SegmentedWidget,
    InfoBar, InfoBarPosition
)
from database import crud
from database.models import Client, Supplier
from ui.db_session import db_session


class EntityDialog(QDialog):
    """Dialogo generico para agregar/editar Clientes y Proveedores."""

    def __init__(self, parent=None, entity=None, entity_type="Cliente"):
        super().__init__(parent)
        self.entity = entity
        self.setWindowTitle(
            f"{'Editar' if entity else 'Agregar'} {entity_type}"
        )
        self.setFixedWidth(460)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 24, 24, 24)

        title = SubtitleLabel(
            f"{'Editar' if entity else 'Nuevo'} {entity_type}", self
        )
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(12)

        self.is_client = entity_type == "Cliente"

        if self.is_client:
            self.property_code_input = LineEdit(self)
            self.property_code_input.setPlaceholderText("Ej: 16-A")
            form.addRow("Código Inmueble:", self.property_code_input)

        self.name_input = LineEdit(self)
        self.name_input.setPlaceholderText("Nombre completo")
        form.addRow("Nombre:", self.name_input)

        self.cedula_input = LineEdit(self)
        self.cedula_input.setPlaceholderText("Ej: V-12345678 / J-87654321-0")
        form.addRow("Cédula / RIF:", self.cedula_input)

        self.phone_input = LineEdit(self)
        self.phone_input.setPlaceholderText("Ej: 0412-1234567")
        form.addRow("Teléfono:", self.phone_input)

        if self.is_client:
            self.email_input = LineEdit(self)
            self.email_input.setPlaceholderText("correo@ejemplo.com")
            form.addRow("Correo:", self.email_input)

            self.keywords_input = LineEdit(self)
            self.keywords_input.setPlaceholderText("Palabras clave separadas por ;")
            form.addRow("Palabras Clave:", self.keywords_input)

        self.accounts_input = LineEdit(self)
        self.accounts_input.setPlaceholderText("Separadas por coma: 0134..., 0105...")
        form.addRow("Cuentas Bancarias:", self.accounts_input)

        layout.addLayout(form)

        if entity:
            if self.is_client:
                self.property_code_input.setText(entity.property_code or "")
                self.email_input.setText(entity.email or "")
                self.keywords_input.setText(entity.keywords or "")
            self.name_input.setText(entity.name or "")
            self.cedula_input.setText(entity.cedula_rif or "")
            self.phone_input.setText(entity.phone or "")
            self.accounts_input.setText(entity.account_numbers or "")

        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        cancel_btn = PushButton("Cancelar", self)
        cancel_btn.clicked.connect(self.reject)
        save_btn = PrimaryPushButton("Guardar", self)
        save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)

    def get_data(self):
        data = {
            "name": self.name_input.text().strip(),
            "cedula_rif": self.cedula_input.text().strip(),
            "phone": self.phone_input.text().strip(),
            "account_numbers": self.accounts_input.text().strip(),
        }
        if self.is_client:
            data["property_code"] = self.property_code_input.text().strip()
            data["email"] = self.email_input.text().strip()
            data["keywords"] = self.keywords_input.text().strip()
        return data


class EntityTable(QWidget):
    """Tabla reutilizable para Clientes o Proveedores."""

    def __init__(self, entity_type: str, condo_id_getter, parent=None):
        super().__init__(parent)
        self.entity_type = entity_type  # "client" or "supplier"
        self.get_condo_id = condo_id_getter
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 12, 0, 0)
        root.setSpacing(12)

        # Toolbar
        toolbar = QHBoxLayout()
        label = "Clientes / Propietarios" if self.entity_type == "client" else "Proveedores"
        self.add_btn = PrimaryPushButton(FluentIcon.ADD, f"Agregar {label[:-1]}", self)
        self.add_btn.clicked.connect(self._on_add)
        toolbar.addWidget(self.add_btn)

        if self.entity_type == "client":
            self.import_btn = PushButton(FluentIcon.DOWNLOAD, "Importar desde archivo", self)
            self.import_btn.clicked.connect(self._on_import)
            toolbar.addWidget(self.import_btn)

            self.template_btn = PushButton(FluentIcon.SAVE, "Descargar Plantilla", self)
            self.template_btn.clicked.connect(self._on_download_template)
            toolbar.addWidget(self.template_btn)

        self.search_box = LineEdit(self)
        self.search_box.setPlaceholderText("Buscar por nombre, cédula, teléfono, inmueble...")
        self.search_box.setMinimumWidth(280)
        self.search_box.textChanged.connect(self._on_search)
        toolbar.addStretch(1)
        toolbar.addWidget(self.search_box)
        root.addLayout(toolbar)

        # Table
        self.table = TableWidget(self)
        if self.entity_type == "client":
            self.table.setColumnCount(9)
            self.table.setHorizontalHeaderLabels(
                ["ID", "Inmueble", "Nombre", "Cédula/RIF", "Teléfono", "Correo", "Palabras Clave", "Cuentas Bancarias", "Acciones"]
            )
        else:
            self.table.setColumnCount(6)
            self.table.setHorizontalHeaderLabels(
                ["ID", "Nombre", "Cédula/RIF", "Teléfono", "Cuentas Bancarias", "Acciones"]
            )
        hh = self.table.horizontalHeader()
        if self.entity_type == "client":
            hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Nombre
            hh.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)  # Cuentas
        else:
            hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        root.addWidget(self.table)

    def _label(self):
        return "Cliente" if self.entity_type == "client" else "Proveedor"

    def load_data(self):
        condo_id = self.get_condo_id()
        if not condo_id:
            self.table.setRowCount(0)
            return
        if self.entity_type == "client":
            entities = crud.get_clients(db_session, condo_id)
        else:
            entities = crud.get_suppliers(db_session, condo_id)
        self._populate(entities)

    def _populate(self, entities):
        self.table.setRowCount(0)
        for row, e in enumerate(entities):
            self.table.insertRow(row)
            if self.entity_type == "client":
                self.table.setItem(row, 0, QTableWidgetItem(str(e.id)))
                self.table.setItem(row, 1, QTableWidgetItem(e.property_code or ""))
                self.table.setItem(row, 2, QTableWidgetItem(e.name or ""))
                self.table.setItem(row, 3, QTableWidgetItem(e.cedula_rif or ""))
                self.table.setItem(row, 4, QTableWidgetItem(e.phone or ""))
                self.table.setItem(row, 5, QTableWidgetItem(e.email or ""))
                self.table.setItem(row, 6, QTableWidgetItem(e.keywords or ""))
                self.table.setItem(row, 7, QTableWidgetItem(e.account_numbers or ""))
                action_col = 8
            else:
                self.table.setItem(row, 0, QTableWidgetItem(str(e.id)))
                self.table.setItem(row, 1, QTableWidgetItem(e.name or ""))
                self.table.setItem(row, 2, QTableWidgetItem(e.cedula_rif or ""))
                self.table.setItem(row, 3, QTableWidgetItem(e.phone or ""))
                self.table.setItem(row, 4, QTableWidgetItem(e.account_numbers or ""))
                action_col = 5

            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(4, 2, 4, 2)
            btn_layout.setSpacing(6)

            edit_btn = PushButton(FluentIcon.EDIT, "", btn_widget)
            edit_btn.setFixedWidth(36)
            edit_btn.setToolTip("Editar")
            edit_btn.clicked.connect(lambda _, ent=e: self._on_edit(ent))

            del_btn = PushButton(FluentIcon.DELETE, "", btn_widget)
            del_btn.setFixedWidth(36)
            del_btn.setToolTip("Eliminar")
            del_btn.clicked.connect(lambda _, ent=e: self._on_delete(ent))

            btn_layout.addWidget(edit_btn)
            btn_layout.addWidget(del_btn)
            btn_layout.addStretch(1)
            self.table.setCellWidget(row, action_col, btn_widget)

    def _on_search(self, text):
        condo_id = self.get_condo_id()
        if not condo_id:
            return
        if not text.strip():
            self.load_data()
            return
        if self.entity_type == "client":
            results = crud.search_clients(db_session, condo_id, text)
        else:
            results = crud.search_suppliers(db_session, condo_id, text)
        self._populate(results)

    def _on_add(self):
        condo_id = self.get_condo_id()
        if not condo_id:
            InfoBar.warning(
                title="Sin condominio activo",
                content="Primero selecciona un condominio en la barra superior.",
                parent=self.window(),
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return
        dlg = EntityDialog(self, entity_type=self._label())
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                return
            if self.entity_type == "client":
                crud.create_client(db_session, condo_id=condo_id, **data)
            else:
                crud.create_supplier(db_session, condo_id=condo_id, **data)
            self.load_data()
            InfoBar.success(
                title="Guardado",
                content=f"{self._label()} agregado correctamente.",
                parent=self.window(),
                position=InfoBarPosition.TOP,
                duration=2500,
            )

    def _on_edit(self, entity):
        dlg = EntityDialog(self, entity=entity, entity_type=self._label())
        if dlg.exec():
            data = dlg.get_data()
            if not data["name"]:
                return
            if self.entity_type == "client":
                crud.update_client(db_session, entity.id, **data)
            else:
                crud.update_supplier(db_session, entity.id, **data)
            self.load_data()

    def _on_delete(self, entity):
        msg = MessageBox(
            "Confirmar eliminación",
            f"¿Eliminar '{entity.name}'?",
            self
        )
        if msg.exec():
            if self.entity_type == "client":
                crud.delete_client(db_session, entity.id)
            else:
                crud.delete_supplier(db_session, entity.id)
            self.load_data()

    def _on_download_template(self):
        """Genera y guarda un archivo .xlsx de plantilla para importar copropietarios."""
        from PyQt6.QtWidgets import QFileDialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Plantilla", "plantilla_copropietarios.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Copropietarios"

        headers = [
            "Código Inmueble", "Nombre Copropietario",
            "Cédula", "Teléfonos (separados por ;)",
            "Correo Electrónico", "Palabras Clave (separadas por ;)"
        ]
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Ejemplo de fila
        example = ["16-A", "Juan Pérez", "V-12345678", "0412-1234567;0414-7654321", "juan@correo.com", "pago movil;transferencia"]
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(italic=True, color="888888")

        # Ajustar ancho de columnas
        widths = [18, 30, 18, 35, 28, 35]
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

        wb.save(path)
        InfoBar.success(
            title="Plantilla descargada",
            content=f"Plantilla guardada en: {path}",
            parent=self.window(),
            position=InfoBarPosition.TOP,
            duration=3000,
        )

    def _on_import(self):
        """Importar copropietarios desde un archivo .xlsx o .txt."""
        condo_id = self.get_condo_id()
        if not condo_id:
            InfoBar.warning(
                title="Sin condominio activo",
                content="Primero selecciona un condominio.",
                parent=self.window(),
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        from PyQt6.QtWidgets import QFileDialog
        import pandas as pd
        from pathlib import Path

        path, _ = QFileDialog.getOpenFileName(
            self, "Importar Copropietarios", "",
            "Archivos soportados (*.xlsx *.xls *.txt *.csv);;Excel (*.xlsx *.xls);;TXT/CSV (*.txt *.csv)"
        )
        if not path:
            return

        try:
            ext = Path(path).suffix.lower()
            if ext in (".xlsx", ".xls"):
                df = pd.read_excel(path, dtype=str)
            else:
                df = pd.read_csv(path, sep=None, engine="python", dtype=str, encoding="utf-8-sig")
        except Exception as e:
            InfoBar.error(
                title="Error al leer archivo",
                content=str(e),
                parent=self.window(),
                position=InfoBarPosition.TOP,
                duration=5000,
            )
            return

        if df.empty or len(df.columns) < 2:
            InfoBar.error(
                title="Formato inválido",
                content="El archivo debe tener al menos 2 columnas (Código Inmueble, Nombre).",
                parent=self.window(),
                position=InfoBarPosition.TOP,
                duration=4000,
            )
            return

        inserted = 0
        skipped = 0
        for _, row in df.iterrows():
            vals = [str(v).strip() if pd.notna(v) else "" for v in row]
            # Extender a 6 columnas si el archivo tiene menos
            while len(vals) < 6:
                vals.append("")

            property_code = vals[0]
            name = vals[1]
            cedula = vals[2]
            phones = vals[3]
            email = vals[4]
            keywords = vals[5]

            if not name or name.lower() in ("nombre copropietario", "nombre"):
                skipped += 1
                continue

            try:
                crud.create_client(
                    db_session, condo_id=condo_id,
                    property_code=property_code,
                    name=name,
                    cedula_rif=cedula,
                    phone=phones,
                    email=email,
                    keywords=keywords,
                    account_numbers=""
                )
                inserted += 1
            except Exception:
                skipped += 1
                continue

        self.load_data()
        InfoBar.success(
            title="Importación completada",
            content=f"{inserted} copropietarios importados. {skipped} filas omitidas.",
            parent=self.window(),
            position=InfoBarPosition.TOP,
            duration=4000,
        )


class EntitiesInterface(QWidget):
    def __init__(self, condo_id_getter, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("EntitiesInterface")
        self.get_condo_id = condo_id_getter
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(36, 28, 36, 28)
        root.setSpacing(16)

        # Header
        header = QHBoxLayout()
        title = TitleLabel("Base de Datos de Entidades", self)
        header.addWidget(title)
        header.addStretch(1)
        root.addLayout(header)

        # Pivot tabs
        self.pivot = SegmentedWidget(self)
        self.pivot.addItem("clients", "Clientes / Propietarios")
        self.pivot.addItem("suppliers", "Proveedores")
        self.pivot.setCurrentItem("clients")
        self.pivot.currentItemChanged.connect(self._on_tab_change)
        root.addWidget(self.pivot)

        # Stacked content
        self.stack = QStackedWidget(self)
        self.clients_table = EntityTable("client", self.get_condo_id, self)
        self.suppliers_table = EntityTable("supplier", self.get_condo_id, self)
        self.stack.addWidget(self.clients_table)
        self.stack.addWidget(self.suppliers_table)
        root.addWidget(self.stack)

    def _on_tab_change(self, key):
        if key == "clients":
            self.stack.setCurrentIndex(0)
        else:
            self.stack.setCurrentIndex(1)

    def refresh(self):
        self.clients_table.load_data()
        self.suppliers_table.load_data()

    def showEvent(self, event):
        super().showEvent(event)
        self.refresh()
